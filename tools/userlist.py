import json
import openpyxl

# 打开一个workbook
wb = openpyxl.load_workbook(filename="./远程成员库.xlsx")

# 获取当前活跃的worksheet,默认就是第一个worksheet
#ws = wb.active

# 当然也可以使用下面的方法

# 获取所有表格(worksheet)的名字
sheets = wb.sheetnames
# 第一个表格的名称
sheet_first = sheets[0]
# 获取特定的worksheet
ws = wb[sheet_first]

# 获取表格所有行和列，两者都是可迭代的
rows = ws.rows
columns = ws.columns

datalist = []
# 迭代所有的行
for i, row in enumerate(rows):
    if row[0].value != "名号" and row[0].value != None:
        data = {
            'id': i,
            'name': row[0].value,
            'avatar': "",
            'roles': row[2].value,
            'skill': row[4].value,
            'skill_ex': row[3].value,
            'city': row[1].value,
            'time': row[6].value,
            'time_ex': row[7].value,
            'projects': row[8].value,
            'mark': row[9].value,
        }
        datalist.append(data)
        # print(data)
        # line = [col.value for col in row]
# print(datalist)

# 通过坐标读取值
# print ws.cell('A1').value  # A表示列,1表示行
# print ws.cell(row=1, column=1).value


# 读写 json 文件
with open('../assets/users.json', 'w', encoding='utf-8') as f:
    json.dump(datalist, f, ensure_ascii=False)
